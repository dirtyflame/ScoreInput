import json
import os
import pandas as pd
from openpyxl import load_workbook
import datetime
from string import capwords
from pathlib import Path


def call_league():
    valid_leagues = load_valid_leagues()

    league = None

    while league not in valid_leagues:

        league = input("Enter league: ").lower()

        if league not in valid_leagues:
            league = input("ERROR, that league cannot be found. Enter league: ").lower()

    league = capwords(league)

    return league


def league_path(league):

    return str(Path("C:/Users/Matt/Desktop/Football Stats/")) + "\\" + league


def log(message):

    current_date = datetime.date.today()
    current_time = datetime.datetime.now().strftime('%H:%M:%S')

    if not os.path.exists("log.txt"):
        print("Error: log.txt does not exist")

    else:
        with open("log.txt", "a") as log_file:
            log_file.write(f"[{current_date}]-[{current_time}] {message}\n")
            log_file.close()


def get_bitmap():  # Doesn't actually work for me but I think it's because I don't have PIL installed
    try:
        directory = os.path.dirname(__file__)
        filename = os.path.join(directory, 'icon.ico')
        return filename
    except Exception as e:
        print(str(e) + ': icon.ico could not be loaded')
        log(f"{e}: icon.ico could not be loaded")


def load_valid_leagues():  # Loads valid leagues

    filename = "leagues.json"

    try:
        fin = open(filename, 'r', encoding='utf-8')
        leagues = json.load(fin)
        fin.close()

        for i in range(len(leagues)):
            leagues[i] = leagues[i].lower()

        return leagues

    except Exception as e:
        print(str(e) + ": " + filename + " could not be loaded")


def load_stats(league):  # Loads stats from file

    filename = os.path.join(league_path(league), "team_stats.json")
    try:
        fin = open(filename, 'r', encoding='utf-8')
        stats = json.load(fin)
        fin.close()
        return stats
    except Exception as e:
        print(str(e) + ": " + filename + " could not be loaded")


def load_recent_form(league):
    filename = os.path.join(league_path(league), "recent_form.json")
    try:
        fin = open(filename, 'r', encoding='utf-8')
        recent_form = json.load(fin)
        fin.close()
        return recent_form
    except Exception as e:
        print(str(e) + ": " + filename + " could not be loaded")


def load_teams(league):  # Loads teams from file
    filename = os.path.join(league_path(league), "teams.json")

    try:
        fin = open(filename, 'r', encoding='utf-8')
        team_names = json.load(fin)
        fin.close()
        return team_names
    except Exception as e:
        return e


def create_dictionary(league):  # Creates a fresh team_stats.json with nested dictionary
    teams = load_teams(league)

    main_dict = {}

    for i in teams:
        main_dict.update({i: {"Scored Home": 0, "Conceded Home": 0, "Scored Away": 0, "Conceded Away": 0,
                              "Played Home": 0, "Played Away": 0, "Recent Form": ""}})

    #  directory = os.path.dirname(__file__)
    #  filename = os.path.join(directory, 'team_stats.json')

    filename = os.path.join(league_path(league), "team_stats.json")

    with open(filename, 'w', encoding='utf-8') as fout:
        json.dump(main_dict, fout)


def create_recent_form_list(league):  # Creates a fresh recent_form.json with nested dictionary

    teams = load_teams(league)

    recent_form = {}

    for i in teams:
        recent_form.update({i: []})

    #  directory = os.path.dirname(__file__)
    #  filename = os.path.join(directory, 'recent_form.json')

    filename = os.path.join(league_path(league), "recent_form.json")

    with open(filename, 'w', encoding='utf-8') as fout:
        json.dump(recent_form, fout)


def stat_update(league, team, attribute, val_update):  # Adds stats to the saved dictionary

    stats_dict = load_stats(league)

    old_value = stats_dict[team].get(attribute)
    new_value = old_value + val_update

    stats_dict[team][attribute] = new_value

    #  directory = os.path.dirname(__file__)
    #  filename = os.path.join(directory, 'team_stats.json')

    filename = os.path.join(league_path(league), "team_stats.json")

    with open(filename, 'w', encoding='utf-8') as fout:
        json.dump(stats_dict, fout)


def add_recent_form_to_stats(league, team):
    stats_dict = load_stats(league)
    form_dict = load_recent_form(league)

    team_form = sum(form_dict[team])
    stats_dict[team]["Recent Form"] = team_form

    #  directory = os.path.dirname(__file__)
    #  filename = os.path.join(directory, 'team_stats.json')

    filename = os.path.join(league_path(league), "team_stats.json")

    with open(filename, 'w', encoding='utf-8') as fout:
        json.dump(stats_dict, fout)


def update_recent_form(league, home_team, away_team, a, b):
    recent_form_dict = load_recent_form(league)

    home_result = None
    away_result = None

    if a > b:
        home_result = 2
        away_result = 0

    elif b > a:
        home_result = 0
        away_result = 2

    else:
        home_result = 1
        away_result = 1

    recent_form_dict[home_team].append(home_result)
    recent_form_dict[away_team].append(away_result)

    if len(recent_form_dict[home_team]) == 7:
        del recent_form_dict[home_team][0]

    if len(recent_form_dict[away_team]) == 7:
        del recent_form_dict[away_team][0]

    #  directory = os.path.dirname(__file__)
    #  filename = os.path.join(directory, 'recent_form.json')

    filename = os.path.join(league_path(league), "recent_form.json")

    with open(filename, 'w', encoding='utf-8') as fout:
        json.dump(recent_form_dict, fout)

    add_recent_form_to_stats(league, home_team)
    add_recent_form_to_stats(league, away_team)


def add_data_to_excel(league):
    jsondata = load_stats(league)
    data = pd.DataFrame(jsondata)

    league_name = capwords(league)

    #  wb = load_workbook(league_path(league) + league + ".xlsx")

    wb = load_workbook(os.path.join(league_path(league), league + ".xlsx"))

    try:
        if wb['Data']:
            std = wb['Data']
            wb.remove(std)

    except KeyError:
        print("The sheet does not exist")

    wb.save(os.path.join(league_path(league), league + ".xlsx"))

    datatoexcel = pd.ExcelWriter(os.path.join(league_path(league), league + ".xlsx"), engine='openpyxl', mode='a')

    data.to_excel(datatoexcel, sheet_name='Data')

    datatoexcel.save()
    log(f"{league_name} stats successfully submitted to Excel.")
